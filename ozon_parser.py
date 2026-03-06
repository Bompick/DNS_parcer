"""
Парсер цен на телевизоры LG с Ozon.ru
Собирает неуценённые ТВ от продавца Ozon и сохраняет в Excel.
"""

import re
import time
import logging
from datetime import datetime
from functools import wraps

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- Конфигурация ---
BASE_URL = "https://www.ozon.ru/category/televizory-15528/tehnika-lg-23969305/"
PARAMS = "seller=0&sorting=price"
MAX_PAGES = 10
SCROLL_PAUSE = 2
PAGE_LOAD_TIMEOUT = 30

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)


# --- Retry декоратор ---
def retry(max_attempts=3, delay=5):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(1, max_attempts + 1):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    logger.warning(f"Попытка {attempt}/{max_attempts} не удалась: {e}")
                    if attempt == max_attempts:
                        raise
                    time.sleep(delay * attempt)
        return wrapper
    return decorator


# --- Настройка браузера ---
def create_driver():
    options = uc.ChromeOptions()
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--lang=ru-RU")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    # Не используем headless — Ozon определяет headless-браузеры

    driver = uc.Chrome(options=options, version_main=145)
    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    driver.implicitly_wait(10)
    return driver


# --- Загрузка страницы ---
def scroll_to_bottom(driver):
    """Прокрутка вниз для подгрузки товаров (lazy-loading)."""
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(20):  # максимум 20 прокруток
        driver.execute_script("window.scrollBy(0, 800);")
        time.sleep(SCROLL_PAUSE)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


@retry(max_attempts=3, delay=5)
def load_page(driver, page_num):
    url = f"{BASE_URL}?{PARAMS}&page={page_num}"
    logger.info(f"Загрузка: {url}")
    driver.get(url)

    # Ждём появления карточек товаров
    WebDriverWait(driver, PAGE_LOAD_TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href*="/product/"]'))
    )
    time.sleep(2)  # Ожидание гидратации React

    scroll_to_bottom(driver)


# --- Фильтрация уценённых товаров ---
def is_utsenennyy(href, text):
    """Проверяет, является ли товар уценённым."""
    # Проверка по URL
    if "utsenennyy" in href.lower() or "utsenenny" in href.lower():
        return True
    # Проверка по тексту названия
    text_lower = text.lower()
    if "уцененн" in text_lower or "уценённ" in text_lower:
        return True
    return False


# --- Извлечение названия модели ---
def extract_model_name(text):
    """Очищает название модели от лишних пометок."""
    name = re.sub(r'[.\s]*[Уу]цененн\w*\s+товар\s*$', '', text, flags=re.IGNORECASE)
    return name.strip()


# --- Извлечение цены из текста ---
def parse_price_text(text):
    """Парсит строку цены вида '21 550 ₽' или '21550' в int."""
    cleaned = re.sub(r'[^\d]', '', text)
    if cleaned:
        try:
            return int(cleaned)
        except ValueError:
            return None
    return None


# --- Парсинг товаров через Selenium (живой DOM) ---
def parse_products_selenium(driver):
    """Парсит товары напрямую через Selenium из живого DOM."""
    products = []
    seen_urls = set()

    # Находим все ссылки на товары
    links = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/product/"]')

    # Собираем уникальные URL товаров с текстом названия
    product_candidates = []
    for link in links:
        href = link.get_attribute("href") or ""
        base_href = href.split("?")[0]
        if base_href in seen_urls:
            continue

        text = link.text.strip()
        if not text:
            continue
        text_lower = text.lower()
        if "lg" not in text_lower:
            continue
        if not any(kw in text_lower for kw in ["телевизор", "televizor", "tv", "тв"]):
            continue

        seen_urls.add(base_href)
        product_candidates.append((link, href, text))

    for link, href, text in product_candidates:
        # Пропускаем уценённые
        if is_utsenennyy(href, text):
            logger.info(f"  [ПРОПУСК] Уценённый: {text[:60]}...")
            continue

        model_name = extract_model_name(text)

        # Поднимаемся по DOM чтобы найти карточку с ценой
        # Используем JS для поиска ближайшего контейнера с ценой
        price = driver.execute_script("""
            let el = arguments[0];
            // Поднимаемся по DOM максимум 15 уровней
            for (let i = 0; i < 15; i++) {
                if (!el.parentElement) break;
                el = el.parentElement;
                let text = el.innerText || '';
                // Ищем цену в формате "число ₽"
                let match = text.match(/(\\d[\\d\\s\\u00a0]*?)\\s*₽/);
                if (match) {
                    return match[1].replace(/\\s/g, '').replace(/\\u00a0/g, '');
                }
            }
            return null;
        """, link)

        parsed_price = None
        if price:
            parsed_price = parse_price_text(price)

        if model_name and parsed_price is not None:
            products.append({
                "model": model_name,
                "price": parsed_price,
            })
            logger.info(f"  [OK] {model_name[:50]}... — {parsed_price:,} ₽")
        else:
            logger.warning(f"  [НЕТ ЦЕНЫ] {model_name[:50]}... (raw={price})")

    return products


# --- Сбор всех товаров по страницам ---
def get_all_products(driver):
    """Обходит все страницы и собирает товары."""
    all_products = []
    seen_models = set()

    for page_num in range(1, MAX_PAGES + 1):
        try:
            load_page(driver, page_num)
        except Exception as e:
            logger.error(f"Не удалось загрузить страницу {page_num}: {e}")
            break

        products = parse_products_selenium(driver)

        if not products:
            logger.info(f"На странице {page_num} товары не найдены, останавливаемся.")
            break

        new_count = 0
        for p in products:
            if p["model"] not in seen_models:
                seen_models.add(p["model"])
                all_products.append(p)
                new_count += 1

        if new_count == 0:
            logger.info(f"Все товары на странице {page_num} — дубликаты, останавливаемся.")
            break

        logger.info(f"Страница {page_num}: +{new_count} новых (всего: {len(all_products)})")
        time.sleep(3)  # Пауза между страницами

    return all_products


import os

# --- Загрузка моделей LG ---
def load_lg_models():
    """
    Loads LG_models.xlsx and returns a dictionary mapping model names to 'LG name'.
    Returns: {model_code: lg_name}
    """
    lg_mapping = {}
    try:
        if not os.path.exists("LG_models.xlsx"):
            logger.warning("LG_models.xlsx not found.")
            return lg_mapping
            
        wb = load_workbook("LG_models.xlsx", read_only=True)
        ws = wb.active
        
        # Headers: model_LG (0), PSI name (1), Name with suffix (2), LG name (3)
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and len(row) > 3 and row[3]:
                model = str(row[0]).strip()
                lg_name = str(row[3]).strip()
                lg_mapping[model] = lg_name
                
        wb.close()
        logger.info(f"Loaded {len(lg_mapping)} LG models from LG_models.xlsx")
    except Exception as e:
        logger.error(f"Error loading LG_models.xlsx: {e}")
        
    return lg_mapping

# --- Сохранение в Excel ---
def save_to_excel(products, filename):
    """Сохраняет список товаров в Excel-файл."""
    # Ensure directory exists
    output_dir = os.path.join("parsing_results", "ozon_parsing")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    filepath = os.path.join(output_dir, filename)
    
    # Load LG models
    lg_mapping = load_lg_models()
    lg_keys = sorted(lg_mapping.keys(), key=len, reverse=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "LG TV Ozon"

    # Стили заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")

    # Заголовки
    headers = ["Название модели", "Lg short name", "Цена (₽)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for i, product in enumerate(products, start=2):
        model_name = product["model"]
        
        # Find LG converter match
        lg_converter = ""
        for key in lg_keys:
            if key in model_name:
                lg_converter = lg_mapping[key]
                break
        
        ws.cell(row=i, column=1, value=model_name)
        ws.cell(row=i, column=2, value=lg_converter)
        
        price_cell = ws.cell(row=i, column=3, value=product["price"])
        price_cell.number_format = '#,##0'
        price_cell.alignment = Alignment(horizontal="right")

    # Ширина колонок
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15

    wb.save(filepath)
    logger.info(f"Сохранено {len(products)} товаров в {filepath}")


# --- Главная функция ---
def main():
    driver = None
    try:
        logger.info("Запуск парсера Ozon — ТВ LG...")
        logger.info("Браузер откроется в видимом режиме.")
        logger.info("Если появится CAPTCHA — решите её вручную, парсер подождёт.")

        driver = create_driver()

        # Даём время на возможную CAPTCHA при первом заходе
        driver.get("https://www.ozon.ru")
        time.sleep(5)

        products = get_all_products(driver)

        if not products:
            logger.warning("Товары не найдены!")
            return

        # Сортировка по цене
        products.sort(key=lambda p: p["price"])

        output_file = f"lg_tv_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        save_to_excel(products, output_file)
        logger.info(f"Готово! Результат: {output_file}")

    except Exception as e:
        logger.error(f"Критическая ошибка: {e}", exc_info=True)
    finally:
        if driver:
            driver.quit()


if __name__ == "__main__":
    main()
