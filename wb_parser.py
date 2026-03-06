"""
Парсер цен на телевизоры LG с Wildberries.ru
Собирает ТВ от продавца Wildberries и сохраняет в Excel.
"""

import re
import time
import logging
import os
from datetime import datetime
from functools import wraps

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- Конфигурация ---
BASE_URL = "https://www.wildberries.ru/catalog/elektronika/tv-audio-foto-video-tehnika/televizory/televizory"
# sort=priceup&page=1&fbrand=5788&fsupplier=-100
PARAMS_TEMPLATE = "sort=priceup&page={}&fbrand=5788&fsupplier=-100"
MAX_PAGES = 10
SCROLL_PAUSE = 2
PAGE_LOAD_TIMEOUT = 30

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)


# --- Retry декоратор ---
def retry(max_attempts=3, delay=5):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(1, max_attempts + 1):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    logger.warning(f"Попытка {attempt}/{max_attempts} не удалась: {e}")
                    if attempt == max_attempts:
                        raise
                    time.sleep(delay * attempt)
        return wrapper
    return decorator


# --- Настройка браузера ---
def create_driver():
    options = uc.ChromeOptions()
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--lang=ru-RU")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    # Не используем headless — WB определяет headless-браузеры

    driver = uc.Chrome(options=options, version_main=145)
    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    driver.implicitly_wait(10)
    return driver


# --- Загрузка страницы ---
def scroll_to_bottom(driver):
    """Прокрутка вниз для подгрузки товаров (lazy-loading)."""
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(20):  # максимум 20 прокруток
        driver.execute_script("window.scrollBy(0, 800);")
        time.sleep(SCROLL_PAUSE)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


@retry(max_attempts=3, delay=5)
def load_page(driver, page_num):
    url = f"{BASE_URL}?{PARAMS_TEMPLATE.format(page_num)}"
    logger.info(f"Загрузка: {url}")
    driver.get(url)

    # Ждём появления карточек товаров
    # WB: .product-card__wrapper or .product-card-list
    WebDriverWait(driver, PAGE_LOAD_TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, '.product-card, .j-card-item, .product-card__wrapper'))
    )
    time.sleep(2)  # Ожидание гидратации

    scroll_to_bottom(driver)


# --- Извлечение названия модели ---
def extract_model_name(text):
    """Очищает название модели от лишних пометок."""
    # WB обычно пишет "Телевизор LG 32LQ63006LA"
    # Удаляем лишнее, если нужно
    return text.strip()


# --- Извлечение цены из текста ---
def parse_price_text(text):
    """Парсит строку цены вида '21 550 ₽' или '21550' в int."""
    cleaned = re.sub(r'[^\d]', '', text)
    if cleaned:
        try:
            return int(cleaned)
        except ValueError:
            return None
    return None


# --- Парсинг товаров через Selenium (живой DOM) ---
def parse_products_selenium(driver):
    """Парсит товары напрямую через Selenium из живого DOM."""
    products = []
    seen_urls = set()

    # Находим все карточки товаров
    # WB selectors might vary
    cards = driver.find_elements(By.CSS_SELECTOR, '.product-card, .j-card-item, .product-card__wrapper')
    
    if not cards:
        # Try finding links directly
        cards = driver.find_elements(By.CSS_SELECTOR, 'a.product-card__link, a.j-card-link')

    logger.info(f"Найдено {len(cards)} карточек/ссылок")

    for card in cards:
        try:
            # Try to get link and text directly if it's an 'a' tag
            if card.tag_name == 'a':
                link_el = card
            else:
                try:
                    link_el = card.find_element(By.CSS_SELECTOR, 'a.product-card__link, a.j-card-link')
                except:
                    continue

            href = link_el.get_attribute("href") or ""
            base_href = href.split("?")[0]
            if base_href in seen_urls:
                continue
            
            # Name extraction
            try:
                # Try finding name element inside card
                name_el = card.find_element(By.CSS_SELECTOR, '.product-card__name, .goods-name, .product-card__brand-name')
                text = name_el.text.strip()
                # Sometimes brand and name are separate
                try:
                    brand_el = card.find_element(By.CSS_SELECTOR, '.product-card__brand')
                    brand_text = brand_el.text.strip()
                    if brand_text and brand_text not in text:
                        text = f"{brand_text} {text}"
                except:
                    pass
            except:
                # Fallback to link text or aria-label
                text = link_el.get_attribute("aria-label") or link_el.text.strip()

            if not text:
                continue
                
            text_lower = text.lower()
            if "lg" not in text_lower:
                continue
            if not any(kw in text_lower for kw in ["телевизор", "televizor", "tv", "тв"]):
                continue

            seen_urls.add(base_href)

            # Price extraction
            price = None
            try:
                # Try finding price element inside card
                # .price__lower-price, .lower-price
                price_el = card.find_element(By.CSS_SELECTOR, '.price__lower-price, .lower-price, .product-card__price ins')
                price_text = price_el.text.strip()
                price = parse_price_text(price_text)
            except:
                # Try JS approach if element not found or text empty
                pass

            if price is None:
                 # Fallback JS to find price near the link
                price_raw = driver.execute_script("""
                    let el = arguments[0];
                    // Go up to card container
                    let card = el.closest('.product-card') || el.closest('.j-card-item') || el.parentElement;
                    if (!card) return null;
                    
                    let priceEl = card.querySelector('.price__lower-price') || card.querySelector('.lower-price') || card.querySelector('ins');
                    return priceEl ? priceEl.innerText : null;
                """, link_el)
                if price_raw:
                    price = parse_price_text(price_raw)

            if text and price is not None:
                products.append({
                    "model": text,
                    "price": price,
                })
                logger.info(f"  [OK] {text[:50]}... — {price:,} ₽")
            else:
                logger.warning(f"  [НЕТ ЦЕНЫ] {text[:50]}...")

        except Exception as e:
            continue

    return products


# --- Сбор всех товаров по страницам ---
def get_all_products(driver):
    """Обходит все страницы и собирает товары."""
    all_products = []
    seen_models = set()

    for page_num in range(1, MAX_PAGES + 1):
        try:
            load_page(driver, page_num)
        except Exception as e:
            logger.error(f"Не удалось загрузить страницу {page_num}: {e}")
            break

        products = parse_products_selenium(driver)

        if not products:
            logger.info(f"На странице {page_num} товары не найдены, останавливаемся.")
            # Save debug html
            with open("debug_wb.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            break

        new_count = 0
        for p in products:
            if p["model"] not in seen_models:
                seen_models.add(p["model"])
                all_products.append(p)
                new_count += 1

        if new_count == 0:
            logger.info(f"Все товары на странице {page_num} — дубликаты, останавливаемся.")
            break

        logger.info(f"Страница {page_num}: +{new_count} новых (всего: {len(all_products)})")
        time.sleep(3)  # Пауза между страницами

    return all_products


# --- Загрузка моделей LG ---
def load_lg_models():
    """
    Loads LG_models.xlsx and returns a dictionary mapping model names to 'LG name'.
    Returns: {model_code: lg_name}
    """
    lg_mapping = {}
    try:
        if not os.path.exists("LG_models.xlsx"):
            logger.warning("LG_models.xlsx not found.")
            return lg_mapping
            
        wb = load_workbook("LG_models.xlsx", read_only=True)
        ws = wb.active
        
        # Headers: model_LG (0), PSI name (1), Name with suffix (2), LG name (3)
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and len(row) > 3 and row[3]:
                model = str(row[0]).strip()
                lg_name = str(row[3]).strip()
                lg_mapping[model] = lg_name
                
        wb.close()
        logger.info(f"Loaded {len(lg_mapping)} LG models from LG_models.xlsx")
    except Exception as e:
        logger.error(f"Error loading LG_models.xlsx: {e}")
        
    return lg_mapping


# --- Сохранение в Excel ---
def save_to_excel(products, filename):
    """Сохраняет список товаров в Excel-файл."""
    # Ensure directory exists
    output_dir = os.path.join("parsing_results", "wb_parsing")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    filepath = os.path.join(output_dir, filename)
    
    # Load LG models
    lg_mapping = load_lg_models()
    lg_keys = sorted(lg_mapping.keys(), key=len, reverse=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "LG TV WB"

    # Стили заголовков
    header_fill = PatternFill(start_color="8A2BE2", end_color="8A2BE2", fill_type="solid") # Purple for WB
    header_font = Font(bold=True, size=12, color="FFFFFF")

    # Заголовки
    headers = ["Название модели", "Lg short name", "Цена (₽)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for i, product in enumerate(products, start=2):
        model_name = product["model"]
        
        # Find LG converter match
        lg_converter = ""
        for key in lg_keys:
            if key in model_name:
                lg_converter = lg_mapping[key]
                break
        
        ws.cell(row=i, column=1, value=model_name)
        ws.cell(row=i, column=2, value=lg_converter)
        
        price_cell = ws.cell(row=i, column=3, value=product["price"])
        price_cell.number_format = '#,##0'
        price_cell.alignment = Alignment(horizontal="right")

    # Ширина колонок
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15

    wb.save(filepath)
    logger.info(f"Сохранено {len(products)} товаров в {filepath}")


# --- Главная функция ---
def main():
    driver = None
    try:
        logger.info("Запуск парсера Wildberries — ТВ LG...")
        logger.info("Браузер откроется в видимом режиме.")
        
        driver = create_driver()

        # Даём время на возможную CAPTCHA
        driver.get("https://www.wildberries.ru")
        time.sleep(5)

        products = get_all_products(driver)

        if not products:
            logger.warning("Товары не найдены!")
            return

        # Сортировка по цене
        products.sort(key=lambda p: p["price"])

        output_file = f"lg_tv_wb_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        save_to_excel(products, output_file)
        logger.info(f"Готово! Результат: {output_file}")

    except Exception as e:
        logger.error(f"Критическая ошибка: {e}", exc_info=True)
    finally:
        if driver:
            driver.quit()


if __name__ == "__main__":
    main()
