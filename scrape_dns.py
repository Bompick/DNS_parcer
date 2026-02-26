import sys
import os
import time
import datetime
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

# Add local packages to path
local_packages = os.path.abspath("./.packages")
if local_packages not in sys.path:
    sys.path.append(local_packages)

try:
    import setuptools
    import undetected_chromedriver as uc
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError as e:
    print(f"Failed to import dependencies: {e}")
    sys.exit(1)

def get_timestamp_filename(brand):
    now = datetime.datetime.now()
    return f"DNS_TV_{brand.upper()}_{now.strftime('%d_%m_%Y_%H_%M')}.xlsx"

def setup_driver():
    options = uc.ChromeOptions()
    # options.add_argument("--headless=new") # Disable headless mode to avoid detection
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--start-maximized") # Start maximized
    options.page_load_strategy = 'eager' 
    
    # Force version 145 since auto-detect is picking 146
    driver = uc.Chrome(options=options, version_main=145)
    return driver

def scroll_gradually(driver):
    total_height = driver.execute_script("return document.body.scrollHeight")
    viewport_height = driver.execute_script("return window.innerHeight")
    current_scroll = 0
    
    while current_scroll < total_height:
        current_scroll += viewport_height / 3 # Scroll slower (1/3 of viewport)
        driver.execute_script(f"window.scrollTo(0, {current_scroll});")
        time.sleep(1) # Longer pause to let elements load
        
        # Update total height in case it grew (infinite scroll behavior)
        new_total_height = driver.execute_script("return document.body.scrollHeight")
        if new_total_height > total_height:
            total_height = new_total_height

def get_brand_selection():
    brands = {
        "1": "lg",
        "2": "samsung",
        "3": "haier",
        "4": "tcl",
        "5": "xiaomi",
        "6": "hisense"
    }
    
    print("\nКакой бренд промониторить?")
    print("1. LG")
    print("2. Samsung")
    print("3. Haier")
    print("4. TCL")
    print("5. Xiaomi")
    print("6. Hisense")
    
    while True:
        choice = input("\nВведите номер бренда (1-6): ").strip()
        if choice in brands:
            return brands[choice]
        print("Неверный выбор. Пожалуйста, введите число от 1 до 6.")

def scrape_dns():
    brand = get_brand_selection()
    print(f"\nSelected brand: {brand.upper()}", flush=True)

    print("Initializing driver...", flush=True)
    driver = None
    try:
        driver = setup_driver()
        print("Driver initialized.", flush=True)
    except Exception as e:
        print(f"Failed to setup driver: {e}", flush=True)
        return

    base_url = f"https://www.dns-shop.ru/catalog/17a8ae4916404e77/televizory/?order=1&stock=now-today-tomorrow-later&brand={brand}"
    
    products_data = []
    page = 1
    
    try:
        while True:
            current_url = f"{base_url}&p={page}" if page > 1 else base_url
            print(f"Scraping page {page}: {current_url}")
            
            driver.get(current_url)
            print("Page loaded. Waiting 10 seconds for manual check/captcha solving...", flush=True)
            time.sleep(10) # Wait for initial load and potential manual captcha solving
            
            # Scroll gradually to trigger lazy loading
            print("Scrolling page to load lazy content...", flush=True)
            scroll_gradually(driver)
            time.sleep(2) # Final wait after scrolling
            
            # Find products
            
            # Wait for at least one price to be visible to ensure lazy loading worked
            try:
                print("Waiting for prices to become visible...", flush=True)
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".product-buy__price, .product-buy__price_active, .product-price__current"))
                )
            except:
                print("Timeout waiting for prices. Proceeding anyway...", flush=True)
            
            # Wait for products to load
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".catalog-product"))
                )
            except:
                print("Timeout waiting for products. Page might be empty or captcha triggered.")
                # Save screenshot for debug
                driver.save_screenshot("debug_screenshot.png")
                print("Saved debug_screenshot.png")

            product_elements = driver.find_elements(By.CSS_SELECTOR, ".catalog-product")
            
            if not product_elements:
                # Try another selector
                product_elements = driver.find_elements(By.XPATH, "//div[contains(@class, 'catalog-product')]")
            
            if not product_elements:
                print("No products found on this page. Stopping.")
                # Print page source for debug
                with open("debug_page.html", "w", encoding="utf-8") as f:
                    f.write(driver.page_source)
                print("Saved debug_page.html")
                break
            
            print(f"Found {len(product_elements)} products on page {page}")
            
            for el in product_elements:
                try:
                    name_el = el.find_element(By.CSS_SELECTOR, ".catalog-product__name")
                    name = name_el.text.strip()
                    
                    # Price can be in different places depending on status
                    # Active price
                    try:
                        # Try multiple price selectors
                        price_selectors = [
                            ".product-buy__price", 
                            ".product-buy__price_active",
                            ".product-price__current",
                            ".product-min-price__current"
                        ]
                        
                        price_el = None
                        for selector in price_selectors:
                            try:
                                price_el = el.find_element(By.CSS_SELECTOR, selector)
                                if price_el and price_el.text.strip():
                                    break
                            except:
                                continue
                        
                        if price_el:
                            # Fix for multiple prices (current + old)
                            # Text might be "22 999 ₽ 25 999 ₽" or "22 999 ₽"
                            # We need to isolate the current price which usually comes first or is separated
                            text = price_el.text.strip()
                            
                            # Split by '₽' to separate prices if multiple exist
                            # Usually format is "Current ₽ Old ₽" or just "Current ₽"
                            parts = text.split('₽')
                            
                            # Take the first part which is typically the current price
                            current_price_part = parts[0]
                            
                            # Extract digits only
                            price_text = ''.join(filter(str.isdigit, current_price_part))
                            price = int(price_text) if price_text else 0
                            is_available = True
                        else:
                            raise Exception("Price element not found")
                    except:
                        price = 0
                        is_available = False
                    
                    # Check for "not available" status
                    # Sometimes it says "Товара нет в наличии" or button is disabled/missing
                    # We can check if buy button exists or text says unavailable
                    try:
                         # This selector might need adjustment based on actual site
                        avail_text = el.find_element(By.CSS_SELECTOR, ".order-avail-wrap").text.lower()
                        if "нет в наличии" in avail_text or "уведомить" in avail_text:
                            is_available = False
                    except:
                        pass # Assume available if we found a price
                        
                    products_data.append({
                        "name": name,
                        "price": price,
                        "available": is_available
                    })
                    
                except Exception as e:
                    print(f"Error parsing product: {e}")
                    continue

            # Check for next page button or just increment page and see if we get products
            # DNS pagination: .pagination-widget__page-link_next
            try:
                next_btn = driver.find_elements(By.CSS_SELECTOR, ".pagination-widget__page-link_next")
                if not next_btn or "disabled" in next_btn[0].get_attribute("class"):
                    print("No next page. Finishing.")
                    break
            except:
                # If we can't find next button, maybe just try next page number until 404 or empty
                pass
            
            page += 1
            # Safety break to avoid infinite loops during testing
            if page > 10: # Increased limit
                print("Reached page limit (safety).")
                break
                
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        
    save_to_excel(products_data, brand)

def save_to_excel(data, brand):
    filename = get_timestamp_filename(brand)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{brand.upper()} TVs"
    
    # Headers
    ws.append(["Название модели", "Цена"])
    
    # Orange fill for unavailable
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    for item in data:
        row = [item["name"], item["price"]]
        ws.append(row)
        
        if not item["available"] or item["price"] == 0:
            # Apply fill to the price cell (column 2)
            cell = ws.cell(row=ws.max_row, column=2)
            cell.fill = orange_fill
            
            # Also apply to name if needed, but prompt said "mark price"
            # cell_name = ws.cell(row=ws.max_row, column=1)
            # cell_name.fill = orange_fill

    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    
    wb.save(filename)
    print(f"Saved {len(data)} items to {filename}")

if __name__ == "__main__":
    print("Starting scrape_dns...", flush=True)
    scrape_dns()
