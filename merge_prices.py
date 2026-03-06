import os
import glob
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def get_latest_file(directory, pattern):
    """
    Finds the latest file matching the pattern in the given directory.
    Returns: (file_path, modification_time) or (None, None)
    """
    full_pattern = os.path.join(directory, pattern)
    files = glob.glob(full_pattern)
    
    if not files:
        return None, None
        
    # Sort by modification time (newest first)
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0], os.path.getmtime(files[0])

def is_today(timestamp):
    """Checks if the timestamp is from today."""
    if timestamp is None:
        return False
    file_date = datetime.datetime.fromtimestamp(timestamp).date()
    today = datetime.datetime.now().date()
    return file_date == today

def read_dns_prices(file_path):
    """
    Reads DNS prices from Excel.
    Expected format: Name, LG converter, Lg short name, Price, ...
    Returns: {lg_short_name: {'name': full_name, 'price': price}}
    """
    prices = {}
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Determine column indices based on headers
        headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
        
        # Default indices (fallback)
        name_idx = 0
        short_name_idx = 2
        price_idx = 3
        
        # Dynamic lookup
        try:
            name_idx = headers.index("Название модели")
            short_name_idx = headers.index("Lg short name")
            price_idx = headers.index("Цена текущая")
        except ValueError:
            print("Warning: Could not find standard headers in DNS file. Using defaults.")
            
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[name_idx] and len(row) > price_idx:
                full_name = str(row[name_idx]).strip()
                short_name = str(row[short_name_idx]).strip() if row[short_name_idx] else ""
                
                # If short name is missing, try to use full name or skip?
                # For now, if short name is missing, we can't join reliably, but let's key by short name
                if not short_name:
                    continue
                    
                try:
                    price = int(row[price_idx])
                except:
                    price = 0
                    
                prices[short_name] = {
                    'name': full_name,
                    'price': price
                }
        wb.close()
    except Exception as e:
        print(f"Error reading DNS file: {e}")
        
    return prices

def read_ozon_prices(file_path):
    """
    Reads Ozon prices from Excel.
    Expected format: Name, Lg short name, Price
    Returns: {lg_short_name: {'name': full_name, 'price': price}}
    """
    prices = {}
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Determine column indices
        headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
        
        name_idx = 0
        short_name_idx = 1
        price_idx = 2
        
        try:
            name_idx = headers.index("Название модели")
            short_name_idx = headers.index("Lg short name")
            price_idx = headers.index("Цена (₽)")
        except ValueError:
            print("Warning: Could not find standard headers in Ozon file. Using defaults.")
            
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[name_idx] and len(row) > price_idx:
                full_name = str(row[name_idx]).strip()
                short_name = str(row[short_name_idx]).strip() if row[short_name_idx] else ""
                
                if not short_name:
                    continue
                    
                try:
                    price = int(row[price_idx])
                except:
                    price = 0
                    
                prices[short_name] = {
                    'name': full_name,
                    'price': price
                }
        wb.close()
    except Exception as e:
        print(f"Error reading Ozon file: {e}")
        
    return prices

def read_wb_prices(file_path):
    """
    Reads WB prices from Excel.
    Expected format: Name, Lg short name, Price
    Returns: {lg_short_name: {'name': full_name, 'price': price}}
    """
    prices = {}
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Determine column indices
        headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
        
        name_idx = 0
        short_name_idx = 1
        price_idx = 2
        
        try:
            name_idx = headers.index("Название модели")
            short_name_idx = headers.index("Lg short name")
            price_idx = headers.index("Цена (₽)")
        except ValueError:
            print("Warning: Could not find standard headers in WB file. Using defaults.")
            
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[name_idx] and len(row) > price_idx:
                full_name = str(row[name_idx]).strip()
                short_name = str(row[short_name_idx]).strip() if row[short_name_idx] else ""
                
                if not short_name:
                    continue
                    
                try:
                    price = int(row[price_idx])
                except:
                    price = 0
                    
                prices[short_name] = {
                    'name': full_name,
                    'price': price
                }
        wb.close()
    except Exception as e:
        print(f"Error reading WB file: {e}")
        
    return prices

def merge_prices():
    # Directories
    dns_dir = "parsing_results"
    ozon_dir = os.path.join("parsing_results", "ozon_parsing")
    wb_dir = os.path.join("parsing_results", "wb_parsing")
    output_dir = os.path.join("parsing_results", "all_sellers")
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    # Find latest files
    dns_file, dns_time = get_latest_file(dns_dir, "DNS_TV_LG_*.xlsx")
    ozon_file, ozon_time = get_latest_file(ozon_dir, "lg_tv_*.xlsx")
    wb_file, wb_time = get_latest_file(wb_dir, "lg_tv_wb_*.xlsx")
    
    print(f"Latest DNS file: {dns_file}")
    print(f"Latest Ozon file: {ozon_file}")
    print(f"Latest WB file: {wb_file}")
    
    if not dns_file:
        print("Warning: Could not find DNS file.")
    if not ozon_file:
        print("Warning: Could not find Ozon file.")
    if not wb_file:
        print("Warning: Could not find WB file.")
        
    # Check if files are from today (optional warning)
    if dns_file and not is_today(dns_time):
        print("Warning: DNS file is not from today.")
    if ozon_file and not is_today(ozon_time):
        print("Warning: Ozon file is not from today.")
    if wb_file and not is_today(wb_time):
        print("Warning: WB file is not from today.")
        
    # Read data
    dns_data = read_dns_prices(dns_file) if dns_file else {}
    ozon_data = read_ozon_prices(ozon_file) if ozon_file else {}
    wb_data = read_wb_prices(wb_file) if wb_file else {}
    
    # Collect all unique short names
    all_models = set(dns_data.keys()) | set(ozon_data.keys()) | set(wb_data.keys())
    sorted_models = sorted(list(all_models))
    
    # Create merged data
    merged_data = []
    for model in sorted_models:
        dns_info = dns_data.get(model, {'name': '-', 'price': '-'})
        ozon_info = ozon_data.get(model, {'name': '-', 'price': '-'})
        wb_info = wb_data.get(model, {'name': '-', 'price': '-'})
        
        merged_data.append({
            'short_name': model,
            'ozon_name': ozon_info['name'],
            'dns_name': dns_info['name'],
            'wb_name': wb_info['name'],
            'ozon_price': ozon_info['price'],
            'dns_price': dns_info['price'],
            'wb_price': wb_info['price']
        })
        
    # Save to Excel
    now = datetime.datetime.now()
    output_filename = f"LG_All_Sellers_{now.strftime('%d_%m_%Y_%H_%M')}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LG All Sellers"
    
    # Headers
    headers = ["LG short name", "Название модели Ozon", "Название модели ДНС", "Название модели WB", "Цена Ozon", "Цена ДНС", "Цена WB"]
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Write data
    for row_idx, item in enumerate(merged_data, start=2):
        ws.cell(row=row_idx, column=1, value=item['short_name'])
        ws.cell(row=row_idx, column=2, value=item['ozon_name'])
        ws.cell(row=row_idx, column=3, value=item['dns_name'])
        ws.cell(row=row_idx, column=4, value=item['wb_name'])
        
        ozon_price_cell = ws.cell(row=row_idx, column=5, value=item['ozon_price'])
        if isinstance(item['ozon_price'], int):
            ozon_price_cell.number_format = '#,##0'
        else:
            ozon_price_cell.alignment = Alignment(horizontal="right")
            
        dns_price_cell = ws.cell(row=row_idx, column=6, value=item['dns_price'])
        if isinstance(item['dns_price'], int):
            dns_price_cell.number_format = '#,##0'
        else:
             dns_price_cell.alignment = Alignment(horizontal="right")

        wb_price_cell = ws.cell(row=row_idx, column=7, value=item['wb_price'])
        if isinstance(item['wb_price'], int):
            wb_price_cell.number_format = '#,##0'
        else:
            wb_price_cell.alignment = Alignment(horizontal="right")
            
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    wb.save(output_path)
    print(f"Merged report saved to {output_path}")

if __name__ == "__main__":
    merge_prices()
